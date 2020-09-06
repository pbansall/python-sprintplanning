from openpyxl import load_workbook
import numpy as np
import math
import pandas as pd
import fire


import warnings
warnings.filterwarnings("ignore")

columns = {}
# Correction of headers for Sprint Sheets
def headerCorrection(sheetname,sheet):
    # moving row 1 to header
	columns[sheetname] = sheet[sheetname].columns.values
	sheet[sheetname].columns = sheet[sheetname].iloc[0]
	sheet[sheetname] = sheet[sheetname].iloc[1:]
	columns[sheetname][0] = sheetname
	

#Changing column names as desired
def changeColumnName(sheetname, oldname, newname,sheet):
	sheet[sheetname].rename(columns = {oldname:newname}, inplace = True)

# Deleting the existing sheet
# Loading new sheet with new data
def loadSheet(bookname,newsheets,sheet):
	#print(sheet[sheetname].loc[0])
	book = load_workbook(bookname)
	for newsheet in newsheets:
		if newsheet in book.sheetnames:
			deleteSheet = book.get_sheet_by_name(newsheet)
			book.remove_sheet(deleteSheet)
		with pd.ExcelWriter(bookname,engine = 'openpyxl') as writer:
			writer.book = book
			#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
			sheet[newsheet].to_excel(writer,sheet_name = newsheet,index = False)

def buildSprint(sprint,df,bookname,sheet):
	headerCorrection(sprint,sheet)
	#changeColumnName(sprint,'Billing & Revenue Mgmt','Billing', sheet)
	cols = list(sheet[sprint].columns)
	#emptying data
	sheet[sprint] = sheet[sprint][0:0]
	newrow = {}
	for index, row in df.iterrows():
		components = str(row['Impacted Product']).split(',')
		components = list(map(str.strip, components))
		components = list(map(str.lower,components))
		for col in cols:
			if col.lower().strip() in components:
				newrow[col] = 'X'
			elif col.lower().strip() == 'feature name' :
				newrow[col] = row['Feature Name']
			else:
				newrow[col] = np.nan
		sheet[sprint] = sheet[sprint].append(newrow,ignore_index = True)
	sheet[sprint].sort_values(["Feature Name"],ascending = True, inplace = True)
	#sheet[sprint].drop_duplicates(subset=None, keep='first', inplace=True)
	sheet[sprint] = sheet[sprint].groupby("Feature Name", as_index = False).first()
	#changeColumnName(sprint,'Billing','Billing & Revenue Mgmt',sheet)
	sheet[sprint] = sheet[sprint].columns.to_frame().T.append(sheet[sprint], ignore_index=True)
	sheet[sprint].columns = columns[sprint]
	loadSheet(bookname,[sprint],sheet)
	
# Prepares each Sprint Sheet format
def prepareSprintSheet(sheet,sprint):
	sheet[sprint] = pd.DataFrame()
	sheet[sprint] = sheet['Sprint'].copy()

def columnMerge(bookname,sprintlist):
	book1 = load_workbook(bookname)
	for sheet in book1.sheetnames:
    		if sheet in sprintlist:
        		updateSheet = book1.get_sheet_by_name(sheet)
        		updateSheet.merge_cells('B1:J1')
        		updateSheet.cell(row =1, column = 2).value = 'Impacted Domain'
        
	book1.save(bookname)

def buildUserStory(sheet, component):
	sheetname = f'USbreakdown_{component}'
	sheet[sheetname] = pd.DataFrame()
	sheet[sheetname]  = sheet['User Story'].copy()
	sheet[sheetname] = sheet[sheetname][0:0]
	sheet[sheetname] = sheet[sheetname].append(sheet['User Story'].loc[ (sheet['User Story']['Impacted Product'].str.split(',').apply(lambda x : component in str(x))) & sheet['User Story']['User Story ID'].apply(lambda x : not str(x).startswith(component.upper()[0:3]))],ignore_index = True)
	sheet[sheetname]['USRef'] = sheet[sheetname]['User Story ID'] 
	sheet[sheetname]['User Story ID']  = f'{component.upper()[0:3]}_'
	accept_criteria = '''Given:
When:
Then: '''
	sheet[sheetname]['Acceptance Criteria'] = accept_criteria
	
	

class run():
	def sprintplanning(self, bookname):
		self.sheet = pd.read_excel(bookname,sheet_name= None)
		#print(sheet.keys())
		sprintlist = []
		for i in set(self.sheet['User Story']['Sprint ']):
			if not math.isnan(i):
				df = self.sheet['User Story'].loc[self.sheet['User Story']['Sprint '] == int(i)].iloc[:,:9]
				sprint = f'Sprint {int(i)}'
				sprintlist.append(sprint)
				print(f'Started {sprint}')
				prepareSprintSheet(self.sheet,sprint)
				buildSprint(sprint,df,bookname,self.sheet)
				print(f'Finished {sprint}')
		columnMerge(bookname, sprintlist)
	def userstory(self,bookname, component):
		self.USsheets = []
		self.sheet = pd.read_excel(bookname,sheet_name= None)
		buildUserStory(self.sheet, component)
		self.USsheets.append(f'USbreakdown_{component}')
		loadSheet(bookname,self.USsheets,self.sheet)

if __name__ == '__main__':
	fire.Fire(run)